import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

SNAPSHOT_PATH = "data/bts_snapshot_full_20260309_165523.json"
OUT_PATH = "data/bts_snapshot_full_20260309_165523.xlsx"


def main() -> None:
    with open(SNAPSHOT_PATH, "r", encoding="utf-8") as f:
        products = json.load(f)

    if not isinstance(products, list):
        raise ValueError("Snapshot file must contain a JSON list")

    columns = [
        "id",
        "ean",
        "name",
        "manufacturer",
        "categories",
        "recommended_price",
        "list_price",
        "stock_list",
        "stock_realtime",
        "price_realtime",
        "availability",
        "leadtime_to_ship",
        "image",
        "flammable",
        "restricted_countries",
        "last_updated",
    ]

    wb = Workbook()
    ws = wb.active
    if ws is None:
        raise RuntimeError("Workbook has no active worksheet")

    ws.title = "products"
    ws.append(columns)

    for p in products:
        ws.append([
            p.get("id"),
            p.get("ean"),
            p.get("name"),
            p.get("manufacturer"),
            p.get("categories"),
            p.get("recommended_price"),
            p.get("list_price"),
            p.get("stock_list"),
            p.get("stock_realtime"),
            p.get("price_realtime"),
            p.get("availability"),
            p.get("leadtime_to_ship"),
            p.get("image"),
            p.get("flammable"),
            json.dumps(p.get("restricted_countries", []), ensure_ascii=False),
            p.get("last_updated"),
        ])

    for col_idx, column_cells in enumerate(ws.iter_cols(), start=1):
        max_length = 0
        column_letter = get_column_letter(col_idx)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    Path(OUT_PATH).parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)

    print(f"Exported {len(products)} products to {OUT_PATH}")


if __name__ == "__main__":
    main()